from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer
from werkzeug.security import check_password_hash, generate_password_hash
import MySQLdb
from flask_wtf import CSRFProtect, FlaskForm, CSRFProtect
from utils.auth import admin_required, terapeuta_required
from werkzeug.utils import secure_filename
import os
from io import BytesIO
from wtforms import StringField
from wtforms.validators import DataRequired
from math import ceil
from reportlab.lib.pagesizes import letter
from reportlab.lib.colors import HexColor
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from io import BytesIO
from flask import Response, flash, redirect, url_for
from flask_login import current_user, login_required
import datetime
from dotenv import load_dotenv
from flask import send_from_directory
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import tempfile
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from threading import Thread

# Carregar variáveis de ambiente do arquivo .env
load_dotenv()

app = Flask(__name__, 
    static_folder='static',
    template_folder='templates'
)
#contexto email
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
mail = Mail(app)

app.config['DB_USER'] = os.getenv('DB_USER')
app.config['DB_NAME'] = os.getenv('DB_NAME')
app.config['DB_PASSWORD'] = os.getenv('DB_PASSWORD')
app.config['DB_HOST'] = os.getenv('DB_HOST')
app.config['DB_PORT'] = os.getenv('DB_PORT')

app.config['SECRET_KEY'] = '1235'  # Mude para uma chave segura
app.config['UPLOAD_FOLDER'] = 'uploads/cartas_recomendacao'
app.config['UPLOAD_FOLDER_COMPROVANTE'] = 'uploads/comprovantes_sessoes'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Limite de 16MB para upload
csrf = CSRFProtect(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Criar pasta de uploads se não existir
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# Adicione estas configurações
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])

class User(UserMixin):
    def __init__(self, id, email, tipo_usuario):
        self.id = id
        self.email = email
        self.tipo_usuario = tipo_usuario
    
    def is_admin(self):
        return self.tipo_usuario == 'admin'
    
    def is_terapeuta(self):
        return self.tipo_usuario == 'terapeuta'
    
    def is_paciente(self):
        return self.tipo_usuario == 'paciente'

def conectar_bd():
    try:
        conn = MySQLdb.connect(
            database= os.getenv('DB_NAME'),
            user= os.getenv('DB_USER'),
            password= os.getenv('DB_PASSWORD'),  # Coloque sua senha aqui se houver
            host= os.getenv('DB_HOST'),
            port= int(os.getenv('DB_PORT'))
        )

        cur = conn.cursor()
        
        # Debug: Mostrar banco atual
        cur.execute("SELECT DATABASE()")
        current_db = cur.fetchone()[0]
        print(f"\nBanco atual: {current_db}")
        
        # Debug: Mostrar todas as tabelas
        cur.execute("""
            SELECT table_schema, table_name 
            FROM information_schema.tables 
            WHERE table_schema NOT IN ('pg_catalog', 'information_schema')
        """)
        tables = cur.fetchall()
        # print("Tabelas disponíveis:", tables)
        
        cur.close()
        return conn
    except Exception as e:
        print(f"Erro ao conectar ao banco de dados: {str(e)}")
        raise

@login_manager.user_loader
def load_user(user_id):
    conn = conectar_bd()
    cur = conn.cursor()
    cur.execute("SELECT id, email, tipo_usuario FROM usuarios WHERE id = %s", (user_id,))
    user = cur.fetchone()
    cur.close()
    conn.close()
    
    if user:
        return User(user[0], user[1], user[2])
    return None

@app.route('/')
def selecao_perfil():
    return render_template('selecao_perfil.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if request.method == 'POST':
        email = request.form['email']
        senha = request.form['senha']
        
        conn = conectar_bd()
        cur = conn.cursor()
        
        try:
            # Primeiro, verifica se o usuário existe e as credenciais estão corretas
            cur.execute("""
                SELECT id, email, senha, tipo_usuario 
                FROM usuarios 
                WHERE email = %s AND tipo_usuario = 'paciente' AND status = true
            """, (email,))
            
            user = cur.fetchone()
            
            if user and check_password_hash(user[2], senha):
                user_obj = User(user[0], user[1], user[3])
                login_user(user_obj)
                
                # Verifica se já existe formulário NAPESE para este usuário
                cur.execute("""
                    SELECT id FROM formulario_napese 
                    WHERE email = %s
                """, (email,))
                
                tem_formulario = cur.fetchone()
                
                if not tem_formulario:
                    # Se não tem formulário, redireciona para o formulário NAPESE
                    flash('Por favor, preencha o formulário de cadastro.', 'info')
                    return redirect(url_for('formulario_napese'))
                else:
                    # Se já tem formulário, redireciona para o dashboard
                    # return redirect(url_for('dashboard'))
                    return render_template('paciente/dashboard.html')
            
            flash('Email ou senha incorretos!', 'error')
            
        except Exception as e:
            print(f"Erro no login: {str(e)}")
            flash('Erro ao realizar login. Por favor, tente novamente.', 'error')
            
        finally:
            cur.close()
            conn.close()
    
    return render_template('login.html', form=form)

@app.route('/login-terapeuta', methods=['GET', 'POST'])
def login_terapeuta():
    form = LoginForm()
    if request.method == 'POST':
        email = request.form['email']
        senha = request.form['senha']
        
        conn = conectar_bd()
        cur = conn.cursor()
        
        try:
            cur.execute("""
                SELECT id, email, senha, tipo_usuario 
                FROM usuarios 
                WHERE email = %s AND tipo_usuario = 'terapeuta' AND status = true
            """, (email,))
            
            user = cur.fetchone()
            
            if user and check_password_hash(user[2], senha):
                user_obj = User(user[0], user[1], user[3])
                login_user(user_obj)
                
                # Verifica se já existe formulário de cadastro para este terapeuta
                cur.execute("""
                    SELECT id FROM terapeuta_napese 
                    WHERE email = %s
                """, (email,))
                
                tem_formulario = cur.fetchone()
                
                if not tem_formulario:
                    # Se não tem formulário, redireciona para o formulário de cadastro de terapeuta
                    flash('Por favor, complete seu cadastro.', 'info')
                    return redirect(url_for('formulario_terapeuta'))
                else:
                    # Se já tem formulário, redireciona para o dashboard
                    return redirect(url_for('dashboard_terapeuta'))
            
            flash('Email ou senha incorretos!', 'error')
            
        except Exception as e:
            print(f"Erro no login: {str(e)}")
            flash('Erro ao realizar login. Por favor, tente novamente.', 'error')
            
        finally:
            cur.close()
            conn.close()
    
    return render_template('login_terapeuta.html', form=form)

@app.route('/login-admin', methods=['GET', 'POST'])
def login_admin():
    form = LoginForm()
    if request.method == 'POST':
        email = request.form['email']
        senha = request.form['senha']
        
        conn = conectar_bd()
        cur = conn.cursor()
        
        try:
            cur.execute("""
                SELECT id, email, senha, tipo_usuario 
                FROM usuarios 
                WHERE email = %s AND tipo_usuario = 'admin' AND status = true
            """, (email,))
            
            user = cur.fetchone()
            
            if user and check_password_hash(user[2], senha):
                user_obj = User(user[0], user[1], user[3])
                login_user(user_obj)
                
                # flash('Login realizado com sucesso!', 'success')
                return redirect(url_for('admin_usuarios'))
            
            flash('Email ou senha incorretos!', 'error')
            
        except Exception as e:
            print(f"Erro no login: {str(e)}")
            flash('Erro ao realizar login. Por favor, tente novamente.', 'error')
            
        finally:
            cur.close()
            conn.close()
    
    return render_template('login_admin.html', form=form)

@app.route('/logout')
def logout():
    try:
        logout_user()  # Desloga o usuário
    except Exception as e:
        print(f"Erro ao realizar logout: {e}")
    return redirect(url_for('selecao_perfil'))  # Redireciona para a página de login

class ResetPasswordForm(FlaskForm):
    senha = StringField('Senha', validators=[DataRequired()])
    confirma_senha = StringField('Confirmar Senha', validators=[DataRequired()])

@app.route('/resetar-senha-terapeuta/<token>', methods=['GET', 'POST'])
def resetar_senha_terapeuta(token):
    form = ResetPasswordForm()  # Instanciar o formulário
    
    try:
        # Verificar token (expira em 1 hora)
        email = serializer.loads(token, salt='recuperar-senha', max_age=3600)
        
        # Verificar se o email existe no banco
        conn = conectar_bd()
        cur = conn.cursor()
        cur.execute("SELECT id FROM usuarios WHERE email = %s AND tipo_usuario = 'terapeuta'", (email,))
        usuario = cur.fetchone()
        cur.close()
        conn.close()
        
        if not usuario:
            flash('Link de recuperação inválido.', 'error')
            return redirect(url_for('esqueci_senha_terapeuta'))
        
        if request.method == 'POST':
            if form.validate_on_submit():
                senha = form.senha.data
                confirma_senha = form.confirma_senha.data
                
                if senha != confirma_senha:
                    flash('As senhas não coincidem.', 'error')
                    return render_template('redefinir_senha_terapeuta.html', form=form)
                
                # Validar senha com regex
                import re
                if not re.match(r'^(?=.*[A-Za-z])(?=.*\d)[A-Za-z\d]{8,}$', senha):
                    flash('A senha deve ter no mínimo 8 caracteres, incluindo letras e números.', 'error')
                    return render_template('redefinir_senha_terapeuta.html', form=form)
                
                try:
                    conn = conectar_bd()
                    cur = conn.cursor()
                    
                    # Atualizar a senha do usuário
                    senha_hash = generate_password_hash(senha)
                    cur.execute("""
                        UPDATE usuarios 
                        SET senha = %s 
                        WHERE email = %s AND tipo_usuario = 'terapeuta'
                    """, (senha_hash, email))
                    
                    conn.commit()
                    flash('Senha alterada com sucesso! Por favor, faça login com sua nova senha.', 'success')
                    return redirect(url_for('login_terapeuta'))
                    
                except Exception as e:
                    print(f"Erro ao atualizar senha: {e}")
                    flash('Erro ao atualizar senha. Por favor, tente novamente.', 'error')
                    return render_template('redefinir_senha_terapeuta.html', form=form)
                    
                finally:
                    if 'cur' in locals():
                        cur.close()
                    if 'conn' in locals():
                        conn.close()
        
        return render_template('redefinir_senha_terapeuta.html', form=form)
        
    except Exception as e:
        print(f"Erro na verificação do token: {e}")
        flash('O link de recuperação é inválido ou expirou.', 'error')
        return redirect(url_for('esqueci_senha_terapeuta'))

@app.route('/esqueci-senha-terapeuta', methods=['GET', 'POST'])
def esqueci_senha_terapeuta():
    form = FlaskForm()  # Para CSRF protection
    
    if request.method == 'POST':
        email = request.form.get('email')
        conn = conectar_bd()
        cur = conn.cursor()
        # Buscar o terapeuta no banco de dados
        cur.execute("SELECT id FROM usuarios WHERE email = %s AND tipo_usuario = 'terapeuta'", (email,))
        terapeuta = cur.fetchone()
        cur.close()
        conn.close()
        if terapeuta:
            # Gerar token seguro
            token = serializer.dumps(email, salt='recuperar-senha')
            
            # Criar link de recuperação
            reset_link = url_for(
                'resetar_senha_terapeuta',
                token=token,
                _external=True
            )
            
            # Criar e enviar email
            msg = Message(
                'Recuperação de Senha - NAPESE',
                sender=app.config['MAIL_USERNAME'],
                recipients=[email]
            )
            
            msg.html = render_template(
                '/email/recuperar_senha.html',
                reset_link=reset_link
            )
            
            mail.send(msg)
            
            flash('Um email com instruções para recuperar sua senha foi enviado.', 'success')
            return redirect(url_for('login_terapeuta'))
        else:
            flash('Email não encontrado em nossa base de dados.', 'error')
            return redirect(url_for('esqueci_senha_terapeuta'))
    
    return render_template('esqueci_senha_terapeuta.html', form=form)

@app.route('/cadastro', methods=['GET', 'POST'])
def cadastro():
    form = FlaskForm()
    if request.method == 'POST':
        try:
            email = request.form['email']
            senha = request.form['senha']
            confirma_senha = request.form['confirma_senha']
            
            if senha != confirma_senha:
                flash('As senhas não coincidem!', 'error')
                return render_template('cadastro.html', form=form)
            
            conn = conectar_bd()
            cur = conn.cursor()
            
            # Verifica se o email já existe
            cur.execute("SELECT id FROM usuarios WHERE email = %s", (email,))
            if cur.fetchone():
                flash('Este email já está cadastrado!', 'error')
                cur.close()
                conn.close()
                return render_template('cadastro.html', form=form)
            
            # Gera o hash da senha
            senha_hash = generate_password_hash(senha)
            
            # Insere o usuário
            cur.execute("""
                INSERT INTO usuarios (email, senha, tipo_usuario, data_criacao, status)
                VALUES (%s, %s, 'paciente', CURRENT_TIMESTAMP, true);
            """, (email, senha_hash))
            
            conn.commit()
            cur.close()
            conn.close()
            
            flash('Cadastro realizado com sucesso! Por favor, faça login para continuar.', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            print(f"Erro ao cadastrar usuário: {str(e)}")
            if 'conn' in locals():
                conn.rollback()
                cur.close()
                conn.close()
            flash('Erro ao realizar cadastro. Por favor, tente novamente.', 'error')
            return render_template('cadastro.html', form=form)
            
    return render_template('cadastro.html', form=form)

@app.route('/dashboard')
@login_required
def dashboard():
    return "Bem-vindo ao Dashboard!"

# Defina uma classe de formulário se ainda não tiver uma
class NapeseForm(FlaskForm):
    nome_completo = StringField('Nome Completo', validators=[DataRequired()])
    cpf = StringField('CPF', validators=[DataRequired()])
    # Adicione todos os outros campos necessários aqui
    # ...

@app.route('/formulario_napese', methods=['GET', 'POST'])
@login_required
def formulario_napese():
    form = NapeseForm()
    
    if request.method == 'POST':
        if form.validate_on_submit():
            try:
                conn = conectar_bd()
                cur = conn.cursor()
                
                # Pega todos os campos do formulário
                dados = {
                    'aprovado': 'pendente',
                    'email': current_user.email,
                    'nome_completo': request.form['nome_completo'],
                    'cpf': request.form['cpf'],
                    'cep': request.form['cep'],
                    'telefones': request.form['telefones'],
                    'data_nascimento': request.form['data_nascimento'],
                    'cidade': request.form['cidade'],
                    'estado': request.form['estado'],
                    'genero': request.form['genero'],
                    'profissao': request.form['profissao'],
                    'preferencia_atendimento': request.form['preferencia_atendimento'],
                    'renda_familiar': request.form['renda_familiar'],
                    'num_dependentes': request.form['num_dependentes'],
                    'sintomas_relevantes': request.form['sintomas_relevantes'],
                    'medicacoes': request.form['medicacoes'],
                    'substancias_psicoativas': request.form['substancias_psicoativas'],
                    'historico_acidentes': request.form['historico_acidentes'],
                    'historico_cirurgias': request.form['historico_cirurgias'],
                    'dores': request.form['dores'],
                    'acompanhamento_psiquiatrico': request.form['acompanhamento_psiquiatrico'],
                    'acompanhamento_psicologico': request.form['acompanhamento_psicologico'],
                    'tecnicas_corporais': request.form['tecnicas_corporais'],
                    'conhece_se': request.form['conhece_se'],
                    'motivo_procura': request.form['motivo_procura'],
                    'vivenciou_trauma': request.form['vivenciou_trauma'],
                    'descricao_evento': request.form['descricao_evento'],
                    'tempo_decorrido': request.form['tempo_decorrido'],
                    'envolveu_violencia': request.form['envolveu_violencia'],
                    'impacto_lembracas': request.form['impacto_lembracas'],
                    'impacto_chateado': request.form['impacto_chateado'],
                    'impacto_evitacao': request.form['impacto_evitacao'],
                    'impacto_evitar_gatilhos': request.form['impacto_evitar_gatilhos'],
                    'impacto_crencas': request.form['impacto_crencas'],
                    'impacto_apreensao': request.form['impacto_apreensao'],
                    'impacto_concentracao': request.form['impacto_concentracao'],
                    'impacto_perda_interesse': request.form['impacto_perda_interesse'],
                }

                # Adicionando campos condicionais
                if 'acidente_violencia' in request.form and request.form['acidente_violencia'].strip():
                    dados['acidente_violencia'] = request.form['acidente_violencia'] == "on"

                if 'causas_naturais' in request.form and request.form['causas_naturais'].strip():
                    dados['causas_naturais'] = request.form['causas_naturais'] == "on"

                if 'nao_se_aplica' in request.form and request.form['nao_se_aplica'].strip():
                    dados['nao_se_aplica'] = request.form['nao_se_aplica'] == "on"

                if 'conheceu_site_trauma' in request.form and request.form['conheceu_site_trauma'].strip():
                    dados['conheceu_site_trauma'] = request.form['conheceu_site_trauma'] == "on"

                if 'conheceu_instagram' in request.form and request.form['conheceu_instagram'].strip():
                    dados['conheceu_instagram'] = request.form['conheceu_instagram'] == "on"

                if 'conheceu_indicacao' in request.form and request.form['conheceu_indicacao'].strip():
                    dados['conheceu_indicacao'] = request.form['conheceu_indicacao'] == "on"

                if 'conheceu_treinamentos' in request.form and request.form['conheceu_treinamentos'].strip():
                    dados['conheceu_treinamentos'] = request.form['conheceu_treinamentos'] == "on"

                if 'conheceu_google' in request.form and request.form['conheceu_google'].strip():
                    dados['conheceu_google'] = request.form['conheceu_google'] == "on"

                if 'conheceu_rede_social' in request.form and request.form['conheceu_rede_social'].strip():
                    dados['conheceu_rede_social'] = request.form['conheceu_rede_social'] == "on"

                if 'conheceu_psicologo' in request.form and request.form['conheceu_psicologo'].strip():
                    dados['conheceu_psicologo'] = request.form['conheceu_psicologo'] == "on"

                if 'conheceu_outro' in request.form and request.form['conheceu_outro'].strip():
                    dados['conheceu_outro'] = request.form['conheceu_outro']

                if 'vivencia_direta' in request.form and request.form['vivencia_direta'].strip():
                    dados['vivencia_direta'] = request.form['vivencia_direta'] == "on"

                if 'vivencia_testemunha' in request.form and request.form['vivencia_testemunha'].strip():
                    dados['vivencia_testemunha'] = request.form['vivencia_testemunha'] == "on"

                if 'vivencia_familiar_amigo' in request.form and request.form['vivencia_familiar_amigo'].strip():
                    dados['vivencia_familiar_amigo'] = request.form['vivencia_familiar_amigo'] == "on"

                if 'vivencia_trabalho' in request.form and request.form['vivencia_trabalho'].strip():
                    dados['vivencia_trabalho'] = request.form['vivencia_trabalho'] == "on"

                if 'vivencia_nenhuma' in request.form and request.form['vivencia_nenhuma'].strip():
                    dados['vivencia_nenhuma'] = request.form['vivencia_nenhuma'] == "on"

                if 'vivencia_outro' in request.form and request.form['vivencia_outro'].strip():
                    dados['vivencia_outro'] = request.form['vivencia_outro']
                
                # Debug: mostrar dados recebidos
                # print("Dados recebidos:", dados)
                
                # Monta a query de inserção dinamicamente
                campos = ', '.join(dados.keys())
                placeholders = ', '.join(['%s'] * len(dados))
                query = f"""
                    INSERT INTO formulario_napese ({campos})
                    VALUES ({placeholders})
                """
                
                cur.execute(query, list(dados.values()))
                conn.commit()
                
                # flash('Cadastro realizado com sucesso! Agora você será redirecionado para o dashboard.', 'success')
                # TALVEZ - criar um def dashboardPaciente para checagens
                return render_template('paciente/dashboard.html')
                
            except Exception as e:
                conn.rollback()
                print(f"Erro ao salvar formulário: {str(e)}")
                if('formulario_napese_cpf_key' in str(e)):
                    flash('Erro ao enviar formulário. CPF já cadastrado', 'error')
                else:
                    flash('Erro ao enviar formulário. Por favor, tente novamente.', 'error')
            finally:
                cur.close()
                conn.close()
                
    return render_template('formulario_napese.html', form=form)

def init_db():
    conn = conectar_bd()
    cur = conn.cursor()
    try:
        # 1. Primeiro, criar os tipos ENUM
        # cur.execute("""
        #     DO $$ 
        #     BEGIN
        #         -- Criar tipo_atendimento
        #         IF NOT EXISTS (SELECT 1 FROM pg_type WHERE typname = 'tipo_atendimento') THEN
        #             CREATE TYPE tipo_atendimento AS ENUM ('PRESENCIAL', 'ONLINE');
        #         END IF;

        #         -- Criar faixa_renda
        #         IF NOT EXISTS (SELECT 1 FROM pg_type WHERE typname = 'faixa_renda') THEN
        #             CREATE TYPE faixa_renda AS ENUM ('SEM_RENDA','ATE_1_SALARIO', '1_A_3_SALARIOS', '3_A_6_SALARIOS', '6_A_9_SALARIOS', '9_A_12_SALARIOS', '12_A_15_SALARIOS', 'ACIMA_15_SALARIOS');
        #         END IF;

        #         -- Criar faixa_dependentes
        #         IF NOT EXISTS (SELECT 1 FROM pg_type WHERE typname = 'faixa_dependentes') THEN
        #             CREATE TYPE faixa_dependentes AS ENUM ('NENHUM', '1_A_2', '3_A_5', 'MAIS_DE_5');
        #         END IF;

        #         -- Criar resposta_sim_nao
        #         IF NOT EXISTS (SELECT 1 FROM pg_type WHERE typname = 'resposta_sim_nao') THEN
        #             CREATE TYPE resposta_sim_nao AS ENUM ('SIM', 'NAO', 'NAO_SEI');
        #         END IF;
        #     END $$;
        # """)
        
        # 2. Criar tabela usuarios
        cur.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INT PRIMARY KEY AUTO_INCREMENT,
                email VARCHAR(255) UNIQUE NOT NULL,
                senha VARCHAR(255) NOT NULL,
                data_criacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status BOOLEAN DEFAULT TRUE,
                tipo_usuario VARCHAR(50) NOT NULL,
                is_admin BOOLEAN DEFAULT FALSE,
                ultimo_login TIMESTAMP
            )
        """)

        # 2. Criar tabela terapeutas_pacientes
        cur.execute("""
            CREATE TABLE IF NOT EXISTS terapeutas_pacientes (
                id INT PRIMARY KEY AUTO_INCREMENT,
                terapeuta_id INTEGER NOT NULL,
                paciente_id INTEGER NOT NULL,
                data_criacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status BOOLEAN DEFAULT TRUE,
                CONSTRAINT fk_terapeutas_pacientes_usuarios_terapeuta FOREIGN KEY (terapeuta_id) REFERENCES usuarios(id),
                CONSTRAINT fk_terapeutas_pacientes_usuarios_paciente FOREIGN KEY (paciente_id) REFERENCES usuarios(id)
            );
        """)

        # 3. Criar tabela formulario_napese
        cur.execute("""
            CREATE TABLE IF NOT EXISTS formulario_napese (
                id INT PRIMARY KEY AUTO_INCREMENT,
                data_cadastro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                aprovado VARCHAR(255) NOT NULL,
                email VARCHAR(255) NOT NULL,
                nome_completo VARCHAR(255) NOT NULL,
                cpf VARCHAR(14) NOT NULL UNIQUE,
                cep VARCHAR(9) NOT NULL,
                telefones VARCHAR(100) NOT NULL,
                cidade VARCHAR(100) NOT NULL,
                estado VARCHAR(2) NOT NULL,
                preferencia_atendimento ENUM('PRESENCIAL', 'ONLINE') NOT NULL,
                renda_familiar ENUM('SEM_RENDA','ATE_1_SALARIO', '1_A_3_SALARIOS', '3_A_6_SALARIOS', '6_A_9_SALARIOS', '9_A_12_SALARIOS', '12_A_15_SALARIOS', 'ACIMA_15_SALARIOS') NOT NULL,
                num_dependentes ENUM('NENHUM', '1_A_2', '3_A_5', 'MAIS_DE_5') NOT NULL,
                data_nascimento DATE NOT NULL,
                genero VARCHAR(50) NOT NULL,
                profissao VARCHAR(100) NOT NULL,
                conheceu_site_trauma BOOLEAN DEFAULT FALSE,
                conheceu_instagram BOOLEAN DEFAULT FALSE,
                conheceu_indicacao BOOLEAN DEFAULT FALSE,
                conheceu_treinamentos BOOLEAN DEFAULT FALSE,
                conheceu_google BOOLEAN DEFAULT FALSE,
                conheceu_rede_social BOOLEAN DEFAULT FALSE,
                conheceu_psicologo BOOLEAN DEFAULT FALSE,
                conheceu_outro VARCHAR(255),
                sintomas_relevantes TEXT NOT NULL,
                medicacoes TEXT,
                substancias_psicoativas TEXT,
                historico_acidentes TEXT,
                historico_cirurgias TEXT,
                dores TEXT,
                acompanhamento_psiquiatrico ENUM('SIM', 'NAO', 'NAO_SEI') NOT NULL,
                acompanhamento_psicologico ENUM('SIM', 'NAO', 'NAO_SEI') NOT NULL,
                tecnicas_corporais TEXT,
                conhece_se BOOLEAN NOT NULL,
                motivo_procura TEXT NOT NULL,
                vivenciou_trauma BOOLEAN NOT NULL,
                descricao_evento TEXT,
                tempo_decorrido VARCHAR(100),
                envolveu_violencia BOOLEAN,
                vivencia_direta BOOLEAN DEFAULT FALSE,
                vivencia_testemunha BOOLEAN DEFAULT FALSE,
                vivencia_familiar_amigo BOOLEAN DEFAULT FALSE,
                vivencia_trabalho BOOLEAN DEFAULT FALSE,
                vivencia_nenhuma BOOLEAN DEFAULT FALSE,
                vivencia_outro TEXT,
                acidente_violencia BOOLEAN DEFAULT FALSE,
                causas_naturais BOOLEAN DEFAULT FALSE,
                nao_se_aplica BOOLEAN DEFAULT FALSE,
                impacto_lembracas INTEGER CHECK (impacto_lembracas BETWEEN 0 AND 4),
                impacto_chateado INTEGER CHECK (impacto_chateado BETWEEN 0 AND 4),
                impacto_evitacao INTEGER CHECK (impacto_evitacao BETWEEN 0 AND 4),
                impacto_evitar_gatilhos INTEGER CHECK (impacto_evitar_gatilhos BETWEEN 0 AND 4),
                impacto_crencas INTEGER CHECK (impacto_crencas BETWEEN 0 AND 4),
                impacto_perda_interesse INTEGER CHECK (impacto_perda_interesse BETWEEN 0 AND 4),
                impacto_apreensao INTEGER CHECK (impacto_apreensao BETWEEN 0 AND 4),
                impacto_concentracao INTEGER CHECK (impacto_concentracao BETWEEN 0 AND 4),
                CONSTRAINT email_valido CHECK (email REGEXP '^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\\.[A-Za-z]{2,}$'),
                CONSTRAINT fk_email FOREIGN KEY (email) REFERENCES usuarios(email)
            )
        """)

        # 4. Criar índices
        try:
            cur.execute("""
                CREATE INDEX idx_formulario_email ON formulario_napese(email);
            """)
        except:
            print('indice idx_formulario_email ja existe')

        try:
            cur.execute("""
                CREATE INDEX idx_formulario_cpf ON formulario_napese(cpf);
            """)
        except:
            print('indice idx_formulario_cpf ja existe')

        try:
            cur.execute("""
                CREATE INDEX idx_formulario_data_cadastro ON formulario_napese(data_cadastro);
            """)
        except:
            print('indice idx_formulario_data_cadastro ja existe')

        # Criar tabela terapeuta_napese
        cur.execute("""
            CREATE TABLE IF NOT EXISTS terapeuta_napese (
                id INT PRIMARY KEY AUTO_INCREMENT,
                nome_completo VARCHAR(255) NOT NULL,
                endereco_consultorio TEXT NULL,
                cidade VARCHAR(100) NOT NULL,
                estado VARCHAR(2) NOT NULL,
                telefone VARCHAR(20) NOT NULL,
                celular VARCHAR(20) NOT NULL,
                email VARCHAR(255) NOT NULL REFERENCES usuarios(email),
                cpf VARCHAR(11) NOT NULL UNIQUE,
                nivel_atual VARCHAR(50) NOT NULL,
                ano_conclusao_avancado2 INTEGER,
                ano_conclusao_sep INTEGER,
                professores_formacao TEXT NOT NULL,
                formacao_academica TEXT NOT NULL,
                participa_grupo_estudo BOOLEAN NOT NULL,
                numero_supervisoes_ultimo_ano INTEGER NOT NULL,
                modalidade VARCHAR(50) NOT NULL,
                faixa_valor_sessao VARCHAR(50) NOT NULL,
                consultorio_acessivel BOOLEAN NOT NULL,
                observacao_acessibilidade TEXT,
                interesse_producao_cientifica BOOLEAN NOT NULL,
                associado_abt BOOLEAN NOT NULL,
                carta_recomendacao_path TEXT,
                comprovante_sessoes_path TEXT,
                sugestoes TEXT,
                concordou_termos BOOLEAN NOT NULL,
                data_cadastro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                status VARCHAR(20) DEFAULT 'pendente'
            )
        """)

        conn.commit()
        print("Banco de dados inicializado com sucesso!")
        
    except Exception as e:
        print(f"Erro ao criar tabelas: {e}")
        conn.rollback()
    finally:
        cur.close()
        conn.close()

class LoginForm(FlaskForm):
    pass  # Não precisamos definir campos aqui se estamos usando HTML puro

class TerapeutaForm(FlaskForm):
    pass  # Não precisamos definir campos aqui se estamos usando HTML puro

# Adicione também uma rota para o dashboard do terapeuta
@app.route('/dashboard-terapeuta')
@login_required
def dashboard_terapeuta():
    return render_template('terapeuta/dashboard.html')

@app.route('/dashboard-paciente')
@login_required
def dashboard_paciente():
    return render_template('paciente/dashboard.html')

# Atualizar a rota existente ou criar se não existir
@app.route('/cadastro-paciente')
def cadastro_paciente():
    return render_template('cadastro_paciente.html')

@app.route('/cadastro-usuario-terapeuta', methods=['GET', 'POST'])
def cadastro_usuario_terapeuta():
    form = TerapeutaForm()
    
    if request.method == 'POST':
        if form.validate_on_submit():
            try:
                if request.form.get('email') and request.form.get('senha'):
                    email = request.form.get('email')
                    tipo_usuario = 'terapeuta'
                    password = request.form.get('senha')

                    # Verificar se o email já existe
                    conn = conectar_bd()
                    cur = conn.cursor()
                    cur.execute("SELECT id FROM usuarios WHERE email = %s AND tipo_usuario = 'terapeuta'", (email,))
                    if cur.fetchone():
                        flash('Este email já está cadastrado no sistema. Por favor, use outro email ou faça login.', 'error')
                        return redirect(url_for('cadastro_usuario_terapeuta'))

                    # Gera o hash da senha
                    hashed_password = generate_password_hash(password)

                    try:
                        # Insere o novo usuário no banco de dados
                        cur.execute("""
                        INSERT INTO usuarios (email, tipo_usuario, senha, status)
                        VALUES (%s, %s, %s, true)
                        """, (email, tipo_usuario, hashed_password))
                        conn.commit()
                            
                        flash('Novo usuário criado com sucesso!', 'success')
                    except Exception as e:
                        print(f"Erro ao criar usuário: {e}")
                        flash('Erro ao criar usuário.', 'error')
                    finally:
                        cur.close()
                        conn.close()

                    return redirect(url_for('login_terapeuta'))
                
            except Exception as e:
                print(f"Erro ao cadastrar terapeuta: {str(e)}")
                if 'conn' in locals():
                    conn.rollback()
                flash('Erro ao realizar cadastro. Por favor, tente novamente.', 'error')
    
    return render_template('cadastro_terapeuta.html', form=form)

# Rotas do Admin
@app.route('/admin/base')
@admin_required
def admin_base():
    return render_template('base.html')

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    return render_template('admin/dashboard.html')

@app.route('/admin-usuarios')
@login_required
def admin_usuarios():
    if not current_user.is_admin():  # Verifica se o usuário é um administrador
        flash('Acesso não autorizado!', 'error')
        return redirect(url_for('dashboard'))
    
    conn = conectar_bd()
    cur = conn.cursor()
    
    try:
        # Buscar usuários ativos (admin, terapeuta e paciente)
        cur.execute("""
            SELECT 
                u.id, 
                COALESCE(t.nome_completo, p.nome_completo, u.email) as nome_completo,
                u.email, 
                u.tipo_usuario, 
                u.data_criacao 
            FROM usuarios u
            LEFT JOIN terapeuta_napese t ON u.email = t.email AND u.tipo_usuario = 'terapeuta'
            LEFT JOIN formulario_napese p ON u.email = p.email AND u.tipo_usuario = 'paciente'
            WHERE u.status = true 
            AND (
                (u.tipo_usuario = 'admin')
                OR (u.tipo_usuario = 'terapeuta' AND t.status = 'aprovado')
                OR (u.tipo_usuario = 'paciente' AND p.aprovado = 'aprovado')
            )
            ORDER BY u.tipo_usuario, nome_completo
        """)
        
        usuarios = cur.fetchall()
        usuarios_formatados = [
            {
                'id': usuario[0],
                'nome_completo': usuario[1],
                'email': usuario[2],
                'tipo_usuario': usuario[3],
                'data_cadastro': usuario[4].strftime("%d/%m/%Y")
            }
            for usuario in usuarios
        ]

        # Buscar pacientes pendentes
        cur.execute("""
            SELECT usuarios.id, usuarios.email, formulario_napese.nome_completo, usuarios.data_criacao 
            FROM usuarios 
            LEFT JOIN formulario_napese ON usuarios.email = formulario_napese.email 
            WHERE usuarios.tipo_usuario = 'paciente' 
            AND formulario_napese.aprovado = 'pendente' 
            AND usuarios.status = true 
            ORDER BY usuarios.data_criacao DESC
        """)
        
        pacientes_pendentes = cur.fetchall()
        pacientes_pendentes_formatados = [
            {
                'id': paciente[0],
                'email': paciente[1],
                'nome_completo': paciente[2],
                'data_cadastro': paciente[3].strftime("%d/%m/%Y %H:%M")
            }
            for paciente in pacientes_pendentes
        ]

        # Buscar terapeutas pendentes
        cur.execute("""
            SELECT usuarios.id, usuarios.email, terapeuta_napese.nome_completo, usuarios.data_criacao 
            FROM usuarios 
            LEFT JOIN terapeuta_napese ON usuarios.email = terapeuta_napese.email 
            WHERE usuarios.tipo_usuario = 'terapeuta' 
            AND terapeuta_napese.status = 'pendente' 
            AND usuarios.status = true 
            ORDER BY usuarios.data_criacao DESC
        """)
        
        terapeutas_pendentes = cur.fetchall()
        terapeutas_pendentes_formatados = [
            {
                'id': terapeuta[0],
                'email': terapeuta[1],
                'nome_completo': terapeuta[2],
                'data_cadastro': terapeuta[3].strftime("%d/%m/%Y %H:%M")
            }
            for terapeuta in terapeutas_pendentes
        ]

        # Buscar vínculos entre terapeutas e pacientes
        cur.execute("""
            SELECT 
                t.nome_completo as nome_terapeuta,
                p.nome_completo as nome_paciente,
                tp.data_criacao as data_vinculo,
                tp.status as status_vinculo
            FROM terapeutas_pacientes tp
            INNER JOIN usuarios ut ON tp.terapeuta_id = ut.id
            INNER JOIN usuarios up ON tp.paciente_id = up.id
            INNER JOIN terapeuta_napese t ON ut.email = t.email
            INNER JOIN formulario_napese p ON up.email = p.email
            ORDER BY t.nome_completo, p.nome_completo
        """)
        
        vinculos = cur.fetchall()
        vinculos_formatados = [
            {
                'nome_terapeuta': vinculo[0],
                'nome_paciente': vinculo[1],
                'data_vinculo': vinculo[2].strftime("%d/%m/%Y"),
                'status_vinculo': 'Ativo' if vinculo[3] else 'Inativo'
            }
            for vinculo in vinculos
        ]

        return render_template(
            'admin/usuarios.html',
            usuarios={'items': usuarios_formatados},
            pacientes_pendentes={'items': pacientes_pendentes_formatados},
            terapeutas_pendentes={'items': terapeutas_pendentes_formatados},
            vinculos={'items': vinculos_formatados},
            form=FlaskForm()
        )

    except Exception as e:
        print(f"Erro ao buscar usuários: {e}")
        flash('Erro ao carregar usuários.', 'error')
        return render_template(
            'admin/usuarios.html',
            usuarios={'items': []},
            pacientes_pendentes={'items': []},
            terapeutas_pendentes={'items': []},
            vinculos={'items': []},
            form=FlaskForm()
        )
    finally:
        cur.close()
        conn.close()

@app.route('/editar-usuario', methods=['POST'])
@login_required
def editar_usuario():
    if not current_user.is_admin():
        flash('Acesso não autorizado!', 'error')
        return redirect(url_for('dashboard'))

    user_id = request.form['user_id']
    email = request.form['email']
    password = request.form['password']

    # Verifica se a senha foi fornecida
    if password:
        # Gera o hash da senha
        hashed_password = generate_password_hash(password)
    else:
        hashed_password = None  # Caso não tenha nova senha, não faz alteração
    
    # Lógica para atualizar o usuário no banco de dados
    try:
        conn = conectar_bd()
        cur = conn.cursor()
        cur.execute("UPDATE usuarios SET email = %s, senha = %s WHERE id = %s", (email, hashed_password, user_id))
        conn.commit()
        flash('Usuário atualizado com sucesso!', 'success')
    except Exception as e:
        flash('Erro ao atualizar o usuário.', 'error')
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('admin_usuarios'))

@app.route('/remover-usuario/<int:user_id>', methods=['POST'])
@login_required
def remover_usuario(user_id):
    form = FlaskForm()

    if not current_user.is_admin():
        flash('Acesso não autorizado!', 'error')
        return redirect(url_for('dashboard'))

    conn = conectar_bd()
    cur = conn.cursor()

    if request.method == 'POST':
        if form.validate_on_submit():
            try:
                # Remover o usuário do banco de dados
                cur.execute("UPDATE usuarios SET status=false WHERE id = %s", (user_id,))
                conn.commit()  # Confirma a exclusão no banco

                flash('Usuário removido com sucesso!', 'success')
            except Exception as e:
                print(f"Erro ao remover o usuário: {e}")
                flash('Erro ao remover o usuário.', 'error')
            finally:
                cur.close()
                conn.close()

    return redirect(url_for('admin_usuarios'))  # Redireciona para a lista de usuários

@app.route('/criar-usuario', methods=['POST'])
@login_required
def criar_usuario():
    if not current_user.is_admin():
        flash('Acesso não autorizado!', 'error')
        return redirect(url_for('dashboard'))

    email = request.form.get('email')
    tipo_usuario = request.form.get('tipo_usuario')
    password = request.form.get('password')

    # Gera o hash da senha
    hashed_password = generate_password_hash(password)

    conn = conectar_bd()
    cur = conn.cursor()
    
    try:
        # Insere o novo usuário no banco de dados
        cur.execute("""
            INSERT INTO usuarios (email, tipo_usuario, senha, status)
            VALUES (%s, %s, %s, true)
        """, (email, tipo_usuario, hashed_password))
        conn.commit()
        
        flash('Novo usuário criado com sucesso!', 'success')
    except Exception as e:
        print(f"Erro ao criar usuário: {e}")
        flash('Erro ao criar usuário.', 'error')
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('admin_usuarios'))

@app.route('/avaliacao_terapeuta/<int:userId>')
@login_required
def avaliacao_terapeuta(userId):
    if not current_user.is_authenticated or not current_user.is_admin():  # Verifica se o usuário está autenticado e é admin
        return jsonify({'error': 'Acesso não autorizado'}), 403

    conn = conectar_bd()
    cur = conn.cursor()

    try:
        # Busca o formulario do terapeuta
        cur.execute("""
            SELECT form.id, form.nome_completo, form.cpf, form.telefone, 
            form.celular, form.email, form.cidade, form.estado, 
            form.endereco_consultorio, form.nivel_atual, form.ano_conclusao_avancado2, form.ano_conclusao_sep,
            form.professores_formacao, form.formacao_academica, form.participa_grupo_estudo, form.numero_supervisoes_ultimo_ano, 
            form.modalidade, form.faixa_valor_sessao, form.consultorio_acessivel, form.observacao_acessibilidade, 
            form.interesse_producao_cientifica, form.associado_abt, form.carta_recomendacao_path, form.sugestoes, 
            form.concordou_termos, form.data_cadastro, form.status, form.comprovante_sessoes_path
            from terapeuta_napese as form inner join usuarios as usu on usu.email = form.email 
            WHERE usu.id = %s and usu.status = true and form.status = 'pendente'
        """, (userId,))

        terapeutas = cur.fetchall()
        # Formata os dados
        terapeuta_formatados = [
            {'id': terapeuta[0], 'nome_completo': terapeuta[1], 'cpf': terapeuta[2], 'telefone': terapeuta[3], 
            'celular': terapeuta[4], 'email': terapeuta[5], 'cidade': terapeuta[6], 'estado': terapeuta[7],
            'endereco_consultorio': terapeuta[8], 'nivel_atual': terapeuta[9], 'ano_conclusao_avancado2': terapeuta[10], 'ano_conclusao_sep': terapeuta[11],
            'professores_formacao': terapeuta[12], 'formacao_academica': terapeuta[13], 'participa_grupo_estudo': terapeuta[14], 'numero_supervisoes_ultimo_ano': terapeuta[15],
            'modalidade': terapeuta[16], 'faixa_valor_sessao': terapeuta[17], 'consultorio_acessivel': terapeuta[18], 'observacao_acessibilidade': terapeuta[19],
            'interesse_producao_cientifica': terapeuta[20], 'associado_abt': terapeuta[21], 'carta_recomendacao_path': terapeuta[22], 'sugestoes': terapeuta[23],
            'concordou_termos': terapeuta[24], 'data_cadastro': terapeuta[25], 'status': terapeuta[26], 'comprovante_sessoes_path': terapeuta[27]
            }
            for terapeuta in terapeutas
        ]

        # print(terapeuta_formatados)
        return jsonify(terapeuta_formatados)
    except Exception as e:
        print(f"Erro ao buscar terapeuta pendente: {e}")
        return jsonify({'error': 'Erro ao buscar terapeuta pendente'}), 500

    finally:
        cur.close()
        conn.close()
        

@app.route('/avaliacao_paciente/<int:userId>')
@login_required
def avaliacao_paciente(userId):
    if not current_user.is_authenticated or not current_user.is_admin():  # Verifica se o usuário está autenticado e é admin
        return jsonify({'error': 'Acesso não autorizado'}), 403

    conn = conectar_bd()
    cur = conn.cursor()

    try:
        # Busca o formulario do paciente
        cur.execute("""
            SELECT form.id, form.nome_completo, form.cpf, form.telefones, 
            form.data_nascimento, form.cep, form.cidade, form.estado, 
            form.genero, form.profissao, form.preferencia_atendimento, form.renda_familiar,
            form.num_dependentes, form.conheceu_site_trauma, form.conheceu_instagram, form.conheceu_indicacao,
            form.conheceu_treinamentos, form.conheceu_google, form.conheceu_rede_social, form.conheceu_psicologo,
            form.conheceu_outro, form.sintomas_relevantes, form.medicacoes, form.substancias_psicoativas,
            form.historico_acidentes, form.historico_cirurgias, form.dores, form.acompanhamento_psiquiatrico, 
            form.acompanhamento_psicologico, form.tecnicas_corporais, form.conhece_se, form.motivo_procura,
            form.vivenciou_trauma, form.descricao_evento, form.tempo_decorrido, form.envolveu_violencia, 
            form.vivencia_direta, form.vivencia_testemunha, form.vivencia_familiar_amigo, form.vivencia_trabalho,
            form.vivencia_nenhuma, form.vivencia_outro, form.impacto_lembracas, form.impacto_evitacao,
            form.impacto_crencas, form.impacto_apreensao, form.acidente_violencia, form.causas_naturais, form.nao_se_aplica,
            form.impacto_concentracao, form.impacto_chateado, form.impacto_evitar_gatilhos, form.impacto_perda_interesse
            from formulario_napese as form inner join usuarios as usu on usu.email = form.email 
            WHERE usu.id = %s and usu.status = true and form.aprovado = 'pendente'
        """, (userId,))
        
        pacientes = cur.fetchall()
        # Formata os dados
        pacientes_formatados = [
            {'id': paciente[0], 'nome_completo': paciente[1], 'cpf': paciente[2], 'telefones': paciente[3], 'data_nascimento': paciente[4], 'cep': paciente[5],
            'cidade': paciente[6], 'estado': paciente[7], 'genero': paciente[8], 'profissao': paciente[9], 'preferencia_atendimento': paciente[10], 'renda_familiar': paciente[11],
            'num_dependentes': paciente[12], 'conheceu_site_trauma': paciente[13], 'conheceu_instagram': paciente[14], 'conheceu_indicacao': paciente[15],
            'conheceu_treinamentos': paciente[16], 'conheceu_google': paciente[17], 'conheceu_rede_social': paciente[18], 'conheceu_psicologo': paciente[19],
            'conheceu_outro': paciente[20], 'sintomas_relevantes': paciente[21], 'medicacoes': paciente[22], 'substancias_psicoativas': paciente[23],
            'historico_acidentes': paciente[24], 'historico_cirurgias': paciente[25], 'dores': paciente[26], 'acompanhamento_psiquiatrico': paciente[27],
            'acompanhamento_psicologico': paciente[28], 'tecnicas_corporais': paciente[29], 'conhece_se': paciente[30], 'motivo_procura': paciente[31],
            'vivenciou_trauma': paciente[32], 'descricao_evento': paciente[33], 'tempo_decorrido': paciente[34], 'envolveu_violencia': paciente[35],
            'vivencia_direta': paciente[36], 'vivencia_testemunha': paciente[37], 'vivencia_familiar_amigo': paciente[38], 'vivencia_trabalho': paciente[39],
            'vivencia_nenhuma': paciente[40], 'vivencia_outro': paciente[41], 'impacto_lembracas': paciente[42], 'impacto_evitacao': paciente[43],
            'impacto_crencas': paciente[44], 'impacto_apreensao': paciente[45], 'acidente_violencia': paciente[46], 'causas_naturais': paciente[47],
            'nao_se_aplica': paciente[48], 'impacto_concentracao': paciente[49], 'impacto_chateado': paciente[50], 'impacto_evitar_gatilhos': paciente[51],
            'impacto_perda_interesse': paciente[52]
            }
            for paciente in pacientes
        ]

        # print(pacientes_formatados)
        return jsonify(pacientes_formatados)

    except Exception as e:
        print(f"Erro ao buscar pacientes pendentes: {e}")
        return jsonify({'error': 'Erro ao buscar pacientes pendentes'}), 500

    finally:
        cur.close()
        conn.close()

@app.route('/get_pacientes/<int:terapeuta_id>')
@login_required
def get_pacientes(terapeuta_id):
    if not current_user.is_authenticated or not current_user.is_admin():  # Verifica se o usuário está autenticado e é admin
        return jsonify({'error': 'Acesso não autorizado'}), 403

    conn = conectar_bd()
    cur = conn.cursor()

    try:
        # Busca os pacientes vinculados ao terapeuta
        cur.execute("""
            select usu_paciente.id, usu_paciente.email from terapeutas_pacientes tp
            inner join usuarios usu_terapeuta on (usu_terapeuta.id = terapeuta_id)
            inner join usuarios usu_paciente on (usu_paciente.id = paciente_id)
            WHERE tp.terapeuta_id = %s and tp.status = true 
        """, (terapeuta_id,))
        
        pacientes = cur.fetchall()

        # Formata os dados
        pacientes_formatados = [
            {'id': paciente[0], 'email': paciente[1]} 
            for paciente in pacientes
        ]

        print(pacientes_formatados)
        return jsonify(pacientes_formatados)

    except Exception as e:
        print(f"Erro ao buscar pacientes vinculados: {e}")
        return jsonify({'error': 'Erro ao buscar pacientes vinculados'}), 500

    finally:
        cur.close()
        conn.close()

@app.route('/pacientes_disponiveis', methods=['GET'])
@login_required
def pacientes_disponiveis():
    if not current_user.is_authenticated or not current_user.is_admin():  # Verifica se o usuário está autenticado e é admin
        flash('Acesso não autorizado!', 'error')
        return redirect(url_for('admin_usuarios'))

    conn = conectar_bd()
    cur = conn.cursor()

    try:
        # Obter IDs dos pacientes já vinculados
        cur.execute("""
            SELECT paciente_id FROM terapeutas_pacientes 
            WHERE status = true
        """)
        vinculados = [row[0] for row in cur.fetchall()]

        # Obter lista de pacientes disponíveis
        if vinculados:
            query = """
                SELECT usuarios.id, usuarios.email FROM usuarios 
                LEFT JOIN formulario_napese ON usuarios.email = formulario_napese.email
                WHERE usuarios.tipo_usuario = 'paciente' 
                AND usuarios.id NOT IN %s 
                AND formulario_napese.aprovado = 'aprovado'
                AND usuarios.status = true
            """
            cur.execute(query, (tuple(vinculados),))
        else:
            query = """
                SELECT usuarios.id, usuarios.email FROM usuarios 
                LEFT JOIN formulario_napese ON usuarios.email = formulario_napese.email
                WHERE usuarios.tipo_usuario = 'paciente' 
                AND formulario_napese.aprovado = 'aprovado'
                AND usuarios.status = true
            """
            cur.execute(query)

        pacientes = cur.fetchall()
        return jsonify([{'id': paciente[0], 'email': paciente[1]} for paciente in pacientes])

    except Exception as e:
        print(f"Erro ao buscar pacientes disponíveis: {e}")
        flash( 'Erro ao buscar pacientes disponíveis!', 'error')
        return redirect(url_for('admin_usuarios'))

    finally:
        cur.close()
        conn.close()

@app.route('/definir_status_terapeuta/<int:terapeuta_napese_id>/<string:aprovado>', methods=['POST'])
@login_required
def definir_status_terapeuta(terapeuta_napese_id, aprovado):
    if not current_user.is_authenticated or not current_user.is_admin():  # Verifica se o usuário está autenticado e é admin
        flash('Acesso não autorizado!', 'error')
        return redirect(url_for('admin_usuarios'))

    conn = conectar_bd()
    cur = conn.cursor()

    try:
        # Obter IDs dos pacientes já vinculados
        cur.execute("""
            UPDATE terapeuta_napese SET status = %s WHERE id = %s
        """, (aprovado,terapeuta_napese_id,))
        conn.commit()
        if aprovado == 'aprovado':
            # flash('Terapeuta aprovado com sucesso!', 'success')
            return redirect(url_for('admin_usuarios'))
        else:
            flash('Terapeuta reprovado!', 'success')
            return redirect(url_for('admin_usuarios'))
    except Exception as e:
        print(f"Erro ao aprovar paciente: {e}")
        return jsonify({
            'status': 'error',
            'message': 'Erro ao aprovar paciente!'
        })
        # flash('Erro ao aprovar paciente!', 'error')
        # return redirect(url_for('admin_usuarios'))
    finally:
        cur.close()
        conn.close()

@app.route('/definir_status_paciente/<int:formulario_napese_id>/<string:aprovado>', methods=['POST'])
@login_required
def definir_status_paciente(formulario_napese_id, aprovado):
    if not current_user.is_authenticated or not current_user.is_admin():  # Verifica se o usuário está autenticado e é admin
        flash('Acesso não autorizado!', 'error')
        return redirect(url_for('admin_usuarios'))

    conn = conectar_bd()
    cur = conn.cursor()

    try:
        # Obter IDs dos pacientes já vinculados
        cur.execute("""
            UPDATE formulario_napese SET aprovado = %s WHERE id = %s
        """, (aprovado,formulario_napese_id,))
        conn.commit()
        if aprovado == 'aprovado':
            # flash('Paciente aprovado com sucesso!', 'success')
            return redirect(url_for('admin_usuarios'))
        else:
            flash('Paciente reprovado!', 'success')
            return redirect(url_for('admin_usuarios'))
        

    except Exception as e:
        print(f"Erro ao aprovar paciente: {e}")
        return jsonify({
            'status': 'error',
            'message': 'Erro ao aprovar paciente!'
        })
        # flash('Erro ao aprovar paciente!', 'error')
        # return redirect(url_for('admin_usuarios'))

    finally:
        cur.close()
        conn.close()

def enviar_email_async(app, msg):
    with app.app_context():
        try:
            mail.send(msg)
            print("Email enviado com sucesso!")
        except Exception as e:
            print(f"Erro ao enviar email: {str(e)}")

def gerar_pdf_e_enviar_email(dados_email, terapeuta_email):
    try:
        # Criamos um contexto da aplicação
        ctx = app.app_context()
        ctx.push()

        try:
            # Gera o PDF usando ReportLab
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            # Estilos personalizados
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                textColor=colors.HexColor('#96d232'),
                alignment=1  # Centralizado
            )

            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=12,
                spaceAfter=15,
                textColor=colors.HexColor('#333333'),
                alignment=0  # Alinhado à esquerda
            )

            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=10,
                textColor=colors.HexColor('#666666')
            )

            # Título principal
            story.append(Paragraph('Formulário do Paciente', title_style))
            story.append(Spacer(1, 20))

            # Dados Pessoais
            story.append(Paragraph('Dados Pessoais', subtitle_style))
            personal_data = [
                [Paragraph('Nome Completo:', normal_style), Paragraph(dados_email['nome_completo'], normal_style)],
                [Paragraph('Email:', normal_style), Paragraph(dados_email['email'], normal_style)],
                [Paragraph('Telefones:', normal_style), Paragraph(dados_email['telefones'], normal_style)],
                [Paragraph('Data de Nascimento:', normal_style), Paragraph(dados_email['data_nascimento'].strftime('%d/%m/%Y') if isinstance(dados_email['data_nascimento'], (datetime.date, datetime.datetime)) else dados_email['data_nascimento'], normal_style)],
                [Paragraph('Gênero:', normal_style), Paragraph(dados_email['genero'], normal_style)],
                [Paragraph('Profissão:', normal_style), Paragraph(dados_email['profissao'], normal_style)],
                [Paragraph('Cidade/Estado:', normal_style), Paragraph(f"{dados_email['cidade']}/{dados_email['estado']}", normal_style)],
                [Paragraph('CEP:', normal_style), Paragraph(dados_email['cep'], normal_style)]
            ]
            t = Table(personal_data, colWidths=[2*inch, 4*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5f5f5')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#333333')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(t)
            story.append(Spacer(1, 20))

            # Informações de Atendimento
            story.append(Paragraph('Informações de Atendimento', subtitle_style))
            atendimento_data = [
                [Paragraph('Preferência de Atendimento:', normal_style), Paragraph(dados_email['preferencia_atendimento'], normal_style)],
                [Paragraph('Renda Familiar:', normal_style), Paragraph(dados_email['renda_familiar'], normal_style)],
                [Paragraph('Número de Dependentes:', normal_style), Paragraph(dados_email['num_dependentes'], normal_style)]
            ]
            t = Table(atendimento_data, colWidths=[2*inch, 4*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5f5f5')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#333333')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(t)
            story.append(Spacer(1, 20))

            # Histórico de Saúde
            story.append(Paragraph('Histórico de Saúde', subtitle_style))
            saude_data = [
                [Paragraph('Sintomas Relevantes:', normal_style), Paragraph(dados_email['sintomas_relevantes'], normal_style)],
                [Paragraph('Medicações:', normal_style), Paragraph(dados_email['medicacoes'], normal_style)],
                [Paragraph('Substâncias Psicoativas:', normal_style), Paragraph(dados_email['substancias_psicoativas'], normal_style)],
                [Paragraph('Histórico de Acidentes:', normal_style), Paragraph(dados_email['historico_acidentes'], normal_style)],
                [Paragraph('Histórico de Cirurgias:', normal_style), Paragraph(dados_email['historico_cirurgias'], normal_style)],
                [Paragraph('Dores:', normal_style), Paragraph(dados_email['dores'], normal_style)],
                [Paragraph('Acompanhamento Psiquiátrico:', normal_style), Paragraph(dados_email['acompanhamento_psiquiatrico'], normal_style)],
                [Paragraph('Acompanhamento Psicológico:', normal_style), Paragraph(dados_email['acompanhamento_psicologico'], normal_style)],
                [Paragraph('Técnicas Corporais:', normal_style), Paragraph(dados_email['tecnicas_corporais'], normal_style)]
            ]
            t = Table(saude_data, colWidths=[2*inch, 4*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5f5f5')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#333333')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(t)
            story.append(Spacer(1, 20))

            # Informações sobre Trauma
            story.append(Paragraph('Informações sobre Trauma', subtitle_style))
            trauma_data = [
                [Paragraph('Vivenciou Trauma:', normal_style), Paragraph('Sim' if dados_email['vivenciou_trauma'] else 'Não', normal_style)],
                [Paragraph('Descrição do Evento:', normal_style), Paragraph(dados_email['descricao_evento'], normal_style)],
                [Paragraph('Tempo Decorrido:', normal_style), Paragraph(dados_email['tempo_decorrido'], normal_style)],
                [Paragraph('Envolveu Violência:', normal_style), Paragraph('Sim' if dados_email['envolveu_violencia'] else 'Não', normal_style)],
                [Paragraph('Motivo da Procura:', normal_style), Paragraph(dados_email['motivo_procura'], normal_style)]
            ]
            t = Table(trauma_data, colWidths=[2*inch, 4*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5f5f5')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#333333')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(t)
            story.append(Spacer(1, 20))

            # Impactos
            story.append(Paragraph('Impactos', subtitle_style))
            impactos_data = [
                [Paragraph('Impacto das Lembranças:', normal_style), Paragraph(str(dados_email['impacto_lembracas']), normal_style)],
                [Paragraph('Impacto da Evitação:', normal_style), Paragraph(str(dados_email['impacto_evitacao']), normal_style)],
                [Paragraph('Impacto nas Crenças:', normal_style), Paragraph(str(dados_email['impacto_crencas']), normal_style)],
                [Paragraph('Impacto na Apreensão:', normal_style), Paragraph(str(dados_email['impacto_apreensao']), normal_style)],
                [Paragraph('Impacto na Concentração:', normal_style), Paragraph(str(dados_email['impacto_concentracao']), normal_style)],
                [Paragraph('Impacto no Humor:', normal_style), Paragraph(str(dados_email['impacto_chateado']), normal_style)],
                [Paragraph('Impacto na Evitação de Gatilhos:', normal_style), Paragraph(str(dados_email['impacto_evitar_gatilhos']), normal_style)],
                [Paragraph('Impacto no Interesse:', normal_style), Paragraph(str(dados_email['impacto_perda_interesse']), normal_style)],
                [Paragraph('Soma dos Impactos:', normal_style), Paragraph(str(dados_email['impacto_soma']), normal_style)]
            ]
            t = Table(impactos_data, colWidths=[2*inch, 4*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f5f5f5')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#333333')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e0e0e0')),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            story.append(t)

            # Rodapé
            story.append(Spacer(1, 30))
            story.append(Paragraph('Documento gerado automaticamente pelo sistema NAPESE', normal_style))
            story.append(Paragraph(f"Data de geração: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", normal_style))

            # Gera o PDF
            doc.build(story)
            pdf_data = buffer.getvalue()
            buffer.close()

            # Prepara e envia o email
            msg = Message(
                'Novo Paciente Vinculado',
                sender=app.config['MAIL_USERNAME'],
                recipients=[terapeuta_email]
            )
            msg.html = render_template('email/vincular_paciente_terapeuta_v3.html', dados=dados_email)
            msg.attach('formulario_paciente.pdf', 'application/pdf', pdf_data)
            mail.send(msg)
            print("Email enviado com sucesso!")

        finally:
            # Sempre remove o contexto da aplicação ao finalizar
            ctx.pop()

    except Exception as e:
        print(f"Erro ao gerar PDF e enviar email: {str(e)}")

@app.route('/vincular_paciente', methods=['POST'])
def vincular_paciente():
    try:
        paciente_id = request.form.get('novo_paciente')
        terapeuta_id = request.form.get('terapeuta_id')
        
        if not paciente_id or not terapeuta_id:
            return jsonify({'status': 'error', 'message': 'IDs do paciente e terapeuta são obrigatórios'}), 400

        conn = conectar_bd()
        cur = conn.cursor()

        # Verifica se o paciente existe
        cur.execute("SELECT email FROM usuarios WHERE id = %s", (paciente_id,))
        paciente = cur.fetchone()
        
        if not paciente:
            return jsonify({'status': 'error', 'message': 'Paciente não encontrado'}), 404
        
        # Verifica se o terapeuta existe
        cur.execute("SELECT email FROM usuarios WHERE id = %s", (terapeuta_id,))
        terapeuta = cur.fetchone()
        
        if not terapeuta:
            return jsonify({'status': 'error', 'message': 'Terapeuta não encontrado'}), 404

        # Verifica se já existe um vínculo ativo
        cur.execute("""
            SELECT id FROM terapeutas_pacientes 
            WHERE paciente_id = %s AND terapeuta_id = %s AND status = 1
        """, (paciente_id, terapeuta_id))
        vinculo_ativo = cur.fetchone()

        if vinculo_ativo:
            return jsonify({'status': 'error', 'message': 'Já existe um vínculo ativo entre este paciente e terapeuta'}), 400

        # Verifica se existe um vínculo inativo
        cur.execute("""
            SELECT id FROM terapeutas_pacientes 
            WHERE paciente_id = %s AND terapeuta_id = %s AND status = 0
        """, (paciente_id, terapeuta_id))
        vinculo_inativo = cur.fetchone()

        if vinculo_inativo:
            # Atualiza o vínculo existente para ativo
            cur.execute("""
                UPDATE terapeutas_pacientes 
                SET status = 1, data_criacao = CURRENT_TIMESTAMP 
                WHERE id = %s
            """, (vinculo_inativo[0],))
        else:
            # Cria um novo vínculo
            cur.execute("""
                INSERT INTO terapeutas_pacientes 
                (paciente_id, terapeuta_id, status, data_criacao) 
                VALUES (%s, %s, 1, CURRENT_TIMESTAMP)
            """, (paciente_id, terapeuta_id))

        # Busca dados completos do paciente para o email
        cur.execute("""
            SELECT p.nome_completo, p.cpf, p.telefones, p.data_nascimento,
                p.cidade, p.estado, p.genero, p.profissao, 
                p.preferencia_atendimento, p.sintomas_relevantes,
                p.medicacoes, p.substancias_psicoativas,
                p.historico_acidentes, p.historico_cirurgias,
                p.dores, p.acompanhamento_psiquiatrico,
                p.acompanhamento_psicologico, p.tecnicas_corporais,
                p.motivo_procura, p.vivenciou_trauma,
                p.descricao_evento, p.tempo_decorrido,
                p.envolveu_violencia, p.impacto_lembracas,
                p.impacto_evitacao, p.impacto_crencas,
                p.impacto_apreensao, p.impacto_concentracao,
                p.impacto_chateado, p.impacto_evitar_gatilhos,
                p.impacto_perda_interesse, p.acidente_violencia,
                p.causas_naturais, p.nao_se_aplica, p.conheceu_site_trauma,
                p.conheceu_instagram, p.conheceu_indicacao, p.conheceu_treinamentos,
                p.conheceu_google, p.conheceu_rede_social, p.conheceu_psicologo,
                p.conheceu_outro, p.cep, p.renda_familiar, p.num_dependentes, p.email
            FROM formulario_napese p
            WHERE p.email = %s
        """, (paciente[0],))
        paciente_dados = cur.fetchone()

        conn.commit()

        # Formata os dados para o email
        dados_email = {
            'nome_completo': paciente_dados[0],
            'cpf': paciente_dados[1],
            'email': paciente_dados[45],
            'telefones': paciente_dados[2],
            'data_nascimento': paciente_dados[3],
            'cidade': paciente_dados[4],
            'estado': paciente_dados[5],
            'genero': paciente_dados[6],
            'profissao': paciente_dados[7],
            'preferencia_atendimento': paciente_dados[8],
            'sintomas_relevantes': paciente_dados[9],
            'medicacoes': paciente_dados[10],
            'substancias_psicoativas': paciente_dados[11],
            'historico_acidentes': paciente_dados[12],
            'historico_cirurgias': paciente_dados[13],
            'dores': paciente_dados[14],
            'acompanhamento_psiquiatrico': paciente_dados[15],
            'acompanhamento_psicologico': paciente_dados[16],
            'tecnicas_corporais': paciente_dados[17],
            'motivo_procura': paciente_dados[18],
            'vivenciou_trauma': paciente_dados[19],
            'descricao_evento': paciente_dados[20],
            'tempo_decorrido': paciente_dados[21],
            'envolveu_violencia': paciente_dados[22],
            'impacto_lembracas': paciente_dados[23],
            'impacto_evitacao': paciente_dados[24],
            'impacto_crencas': paciente_dados[25],
            'impacto_apreensao': paciente_dados[26],
            'impacto_concentracao': paciente_dados[27],
            'impacto_chateado': paciente_dados[28],
            'impacto_evitar_gatilhos': paciente_dados[29],
            'impacto_perda_interesse': paciente_dados[30],
            'acidente_violencia': paciente_dados[31],
            'causas_naturais': paciente_dados[32],
            'nao_se_aplica': paciente_dados[33],
            'conheceu_site_trauma': paciente_dados[34],
            'conheceu_instagram': paciente_dados[35],
            'conheceu_indicacao': paciente_dados[36],
            'conheceu_treinamentos': paciente_dados[37],
            'conheceu_google': paciente_dados[38],
            'conheceu_rede_social': paciente_dados[39],
            'conheceu_psicologo': paciente_dados[40],
            'conheceu_outro': paciente_dados[41],
            'cep': paciente_dados[42],
            'renda_familiar': paciente_dados[43],
            'num_dependentes': paciente_dados[44],
            'impacto_soma': int(paciente_dados[23]) + int(paciente_dados[24]) + int(paciente_dados[25]) + 
                           int(paciente_dados[26]) + int(paciente_dados[27]) + int(paciente_dados[28]) + 
                           int(paciente_dados[29]) + int(paciente_dados[30])
        }

        # Inicia o processo assíncrono de geração do PDF e envio do email
        try:
            thread = Thread(target=gerar_pdf_e_enviar_email, args=(dados_email, terapeuta[0]))
            thread.daemon = True  # Garante que a thread será encerrada quando o programa principal terminar
            thread.start()
            print("Thread de envio de email iniciada")
        except Exception as e:
            print(f"Erro ao iniciar thread de email: {str(e)}")

        return jsonify({'status': 'success', 'message': 'Paciente vinculado com sucesso! O email será enviado em breve.'})

    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
        return jsonify({'status': 'error', 'message': f'Erro ao vincular paciente: {str(e)}'}), 500
    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()

@app.route('/remover_vinculo', methods=['POST'])
@login_required
def remover_vinculo():
    terapeuta_id = request.form.get('terapeuta_id')
    paciente_id = request.form.get('paciente_id')

    print(f"terapeuta_id: {terapeuta_id}, paciente_id: {paciente_id}")

    if not terapeuta_id or not paciente_id:
        flash('Parâmetros inválidos', 'error')
        return redirect(url_for('admin_usuarios'))

    conn = conectar_bd()
    cur = conn.cursor()

    try:
        # Verificar se o terapeuta e paciente estão vinculados
        cur.execute(
            "SELECT 1 FROM terapeutas_pacientes WHERE terapeuta_id = %s AND paciente_id = %s",
            (terapeuta_id, paciente_id)
        )
        if not cur.fetchone():
            flash('Vínculo não encontrado', 'error')
            return redirect(url_for('admin_usuarios'))

        # Remover o vínculo
        cur.execute(
            "UPDATE terapeutas_pacientes SET status=false WHERE terapeuta_id = %s AND paciente_id = %s",
            (terapeuta_id, paciente_id)
        )
        conn.commit()

        # flash('Vínculo removido com sucesso', 'success')
        return redirect(url_for('admin_usuarios'))

    except Exception as e:
        print(f"Erro ao remover vínculo: {e}")
        conn.rollback()
        flash('Erro ao remover vínculo', 'error')
        return redirect(url_for('admin_usuarios'))

    finally:
        cur.close()
        conn.close()

@app.route('/gerar_pdf/<user_email>')
@login_required
def gerar_pdf(user_email):
    # Verificar se o usuário tem permissão para acessar
    if not current_user.is_authenticated or not current_user.is_admin():
        flash('Acesso não autorizado!', 'error')
        return redirect(url_for('admin_usuarios'))

    conn = conectar_bd()
    cur = conn.cursor()

    try:
        # Consulta as informações do usuário pelo email
        cur.execute("""
            SELECT *
            FROM formulario_napese 
            WHERE email = %s
        """, (user_email,))
        paciente = cur.fetchone()

        if not paciente:
            flash('Dados do paciente não encontrados!', 'error')
            return redirect(url_for('admin_usuarios'))

        # Nome das colunas da tabela
        colunas = [
            "ID", "Data de Cadastro", "Email", "Nome Completo", "CPF", "CEP", "Telefones", "Cidade", "Estado",
            "Preferência de Atendimento", "Renda Familiar", "Número de Dependentes", "Data de Nascimento", "Gênero",
            "Profissão", "Conheceu pelo Site/Trauma", "Conheceu pelo Instagram", "Conheceu por Indicação",
            "Conheceu pelos Treinamentos", "Conheceu pelo Google", "Conheceu por Redes Sociais", "Conheceu por Psicólogo",
            "Conheceu por Outro Meio", "Sintomas Relevantes", "Medicações", "Substâncias Psicoativas", "Histórico de Acidentes",
            "Histórico de Cirurgias", "Dores", "Acompanhamento Psiquiátrico", "Acompanhamento Psicológico", "Técnicas Corporais",
            "Se Conhece", "Motivo da Procura", "Vivenciou Trauma", "Descrição do Evento", "Tempo Decorrido", "Envolveu Violência",
            "Vivência Direta", "Vivência como Testemunha", "Vivência de Familiar ou Amigo", "Vivência no Trabalho",
            "Sem Vivência", "Outra Vivência", "Impacto das Lembranças", "Impacto da Evitação", "Impacto nas Crenças",
            "Impacto na Apreensão"
        ]

        # Separar os dados em dois blocos
        dados_paciente = [
            (colunas[i], paciente[i]) for i in range(15)
        ]

        dados_internacao = [
            (colunas[i], paciente[i]) for i in range(15, len(colunas))
        ]

        # Ajustar formatação dos dados
        def formatar_valor(valor):
            if isinstance(valor, bool):
                return "Sim" if valor else "Não"
            elif isinstance(valor, (datetime.date, datetime.datetime)):
                return valor.strftime("%d/%m/%Y")
            return valor

        dados_paciente = [(coluna, formatar_valor(valor)) for coluna, valor in dados_paciente]
        dados_internacao = [(coluna, formatar_valor(valor)) for coluna, valor in dados_internacao]

        # Criar o PDF em memória
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)

        # Configurações gerais
        largura, altura = letter

        # Cabeçalho
        c.setFillColor(HexColor("#96d232"))  # Cor laranja
        c.rect(0, altura - 70, largura, 70, stroke=0, fill=1)  # Retângulo do cabeçalho
        c.setFillColor(colors.white)
        c.setFont("Helvetica-Bold", 20)
        c.drawCentredString(largura / 2, altura - 50, "Relatório do Paciente")

        # Renderizar seção "Dados do Paciente"
        y = altura - 100
        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(colors.black)
        c.drawString(50, y, "Dados do Paciente")
        y -= 10
        c.line(50, y, largura - 50, y)
        y -= 30

        c.setFont("Helvetica", 12)
        for coluna, valor in dados_paciente:
            if y < 50:  # Quebra de página
                c.showPage()
                c.setFont("Helvetica-Bold", 14)
                c.setFillColor(colors.black)
                c.drawString(50, altura - 100, "Dados do Paciente")
                y = altura - 130
                c.setFont("Helvetica", 12)  # Resetar fonte após o título
            c.drawString(50, y, f"{coluna}: {valor}")
            y -= 20

        # Renderizar seção "Dados da Internação"
        if y < 100:
            c.showPage()
            y = altura - 100

        y -= 10
        c.setFont("Helvetica-Bold", 14)
        c.setFillColor(colors.black)
        c.drawString(50, y, "Outras Informações")
        y -= 10
        c.line(50, y, largura - 50, y)
        y -= 30

        c.setFont("Helvetica", 12)
        for coluna, valor in dados_internacao:
            if y < 50:  # Quebra de página
                c.showPage()
                c.setFont("Helvetica-Bold", 14)
                c.setFillColor(colors.black)
                # c.drawString(50, altura - 100, "Dados da Internação")
                y = altura - 50
                c.setFont("Helvetica", 12)  # Resetar fonte após o título
            c.drawString(50, y, f"{coluna}: {valor}")
            y -= 20

        # Finalizar o PDF
        c.save()
        buffer.seek(0)

        # Retornar o PDF como resposta
        return Response(
            buffer,
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment;filename=relatorio_paciente.pdf'}
        )

    except Exception as e:
        print(f"Erro ao gerar PDF: {e}")
        flash('Erro ao gerar PDF!', 'error')
        return redirect(url_for('admin_usuarios'))

    finally:
        cur.close()
        conn.close()

# Rotas do Terapeuta
@app.route('/terapeuta/dashboard')
@terapeuta_required
def terapeuta_dashboard():
    return render_template('terapeuta/dashboard.html')

@app.route('/terapeuta/meus-pacientes')
@terapeuta_required
def terapeuta_pacientes():
    return render_template('terapeuta/pacientes.html')

# Rotas do Paciente (não precisam de decorator especial, apenas login_required)
@app.route('/paciente/perfil')
@login_required
def paciente_perfil():
    return render_template('paciente/perfil.html')

def criar_usuarios_teste():
    print("Criando usuários de teste...")
    conn = conectar_bd()
    cur = conn.cursor()
    try:
        # Criar admin
        cur.execute("""
            INSERT INTO usuarios (email, senha, tipo_usuario, status)
            VALUES (%s, %s, 'admin', true)
            ON DUPLICATE KEY UPDATE tipo_usuario = 'admin'
        """, ('admin@teste.com', generate_password_hash('admin123')))

        # Criar terapeuta
        cur.execute("""
            INSERT INTO usuarios (email, senha, tipo_usuario, status)
            VALUES (%s, %s, 'terapeuta', true)
            ON DUPLICATE KEY UPDATE tipo_usuario = 'terapeuta'
        """, ('terapeuta@teste.com', generate_password_hash('terapeuta123')))

        # Criar paciente
        cur.execute("""
            INSERT INTO usuarios (email, senha, tipo_usuario, status)
            VALUES (%s, %s, 'paciente', true)
            ON DUPLICATE KEY UPDATE tipo_usuario = 'paciente'
        """, ('paciente@teste.com', generate_password_hash('paciente123')))

        conn.commit()
        print("Usuários de teste criados com sucesso!")
        
    except Exception as e:
        print(f"Erro ao criar usuários: {str(e)}")
        conn.rollback()
    finally:
        cur.close()
        conn.close()

# Configurações para upload de arquivos
UPLOAD_FOLDER = 'uploads/cartas_recomendacao'
UPLOAD_FOLDER_COMPROVANTE = 'uploads/comprovantes_sessoes'
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/formulario-terapeuta', methods=['GET', 'POST'])
@login_required
def formulario_terapeuta():
    form = FlaskForm()
    
    # Adicione este print para debug
    print("Method:", request.method)
    if request.method == 'POST':
        # print("Form data:", request.form)  # Debug dos dados do formulário
        # print("Files:", request.files)     # Debug dos arquivos
        
        try:
            # Verifica se já existe um formulário preenchido
            conn = conectar_bd()
            cur = conn.cursor()
            
            # print(f"Verificando formulário existente para email: {current_user.email}") # Debug log
            cur.execute("""
                SELECT id 
                FROM terapeuta_napese 
                WHERE email = %s
            """, (current_user.email,))

            
            if cur.fetchone():
                # print("Formulário já preenchido.") # Debug log
                flash('Você já preencheu o formulário anteriormente.', 'info')
                return redirect(url_for('dashboard_terapeuta'))
            
            # Processar upload da carta de recomendação
            carta_path = None
            if 'carta_recomendacao' in request.files:
                arquivo = request.files['carta_recomendacao']
                if arquivo and arquivo.filename and allowed_file(arquivo.filename):
                    filename = secure_filename(f"{current_user.email}_{arquivo.filename}")
                    if not os.path.exists(UPLOAD_FOLDER):
                        os.makedirs(UPLOAD_FOLDER)
                    arquivo.save(os.path.join(UPLOAD_FOLDER, filename))
                    carta_path = filename
            comprovante_path = None
            if 'comprovante_sessoes' in request.files:
                arquivo_comprovante = request.files['comprovante_sessoes']
                if arquivo_comprovante and arquivo_comprovante.filename and allowed_file(arquivo_comprovante.filename):
                    comprovante_filename = secure_filename(f"{current_user.email}_compr_{arquivo_comprovante.filename}")
                    if not os.path.exists(UPLOAD_FOLDER_COMPROVANTE):
                        os.makedirs(UPLOAD_FOLDER_COMPROVANTE)
                    arquivo_comprovante.save(os.path.join(UPLOAD_FOLDER_COMPROVANTE,comprovante_filename))
                    comprovante_path = comprovante_filename
            endereco_consultorio = ''
            if 'endereco_consultorio' in request.form:
                endereco_consultorio = request.form['endereco_consultorio']
            consultorio_acessivel = 0
            if 'consultorio_acessivel' in request.form:
                if request.form['consultorio_acessivel'] != '0':
                    consultorio_acessivel = 1
            # print("Carta path:", carta_path) # Debug log
            # Coletar dados do formulário
            try:
                numero_supervisoes = request.form.get('numero_supervisoes_ultimo_ano', '0')
                numero_supervisoes = int(numero_supervisoes) if numero_supervisoes.strip() else 0
            except (ValueError, TypeError):
                numero_supervisoes = 0

            dados = {
                'nome_completo': request.form['nome_completo'],
                'endereco_consultorio': endereco_consultorio,
                'cidade': request.form['cidade'],
                'estado': request.form['estado'],
                'telefone': request.form['telefone'],
                'celular': request.form['celular'],
                'email': current_user.email,
                'cpf': request.form['cpf'].replace('.', '').replace('-', ''),
                'nivel_atual': request.form['nivel_atual'],
                # descomentar após adicionar os campos no html
                'ano_conclusao_avancado2': int(request.form.get('ano_conclusao_avancado2',0)),
                'ano_conclusao_sep': int(request.form.get('ano_conclusao_sep',0)),
                'professores_formacao': request.form['professores_formacao'],
                'formacao_academica': request.form['formacao_academica'],
                'participa_grupo_estudo': request.form.get('participa_grupo_estudo') == '1',
                'interesse_producao_cientifica': request.form.get('interesse_producao_cientifica') == '1',
                'associado_abt': request.form.get('associado_abt') == '1',
                'carta_recomendacao_path': carta_path,
                'comprovante_sessoes_path': comprovante_path,
                'sugestoes': request.form.get('sugestoes', ''),

                'numero_supervisoes_ultimo_ano': numero_supervisoes,
                'modalidade': request.form['modalidade'],
                # 'faixa_valor_sessao': request.form['faixa_valor_sessao'],
                'consultorio_acessivel': request.form.get('consultorio_acessivel') == '1',
                'observacao_acessibilidade': request.form.get('observacao_acessibilidade', ''),
                # apagar após adicionar os campos no html
                'faixa_valor_sessao': '-',
                'concordou_termos': 1
            }

            # print("Dados coletados:", dados) # Debug log

            # Construir query de inserção
            campos = ', '.join(dados.keys())
            # placeholders = ', '.join([get_pg_placeholder(value) for value in dados.values()])
            placeholders = ', '.join(['%s'] * len(dados))
            query = f"""
                INSERT INTO terapeuta_napese ({campos})
                VALUES ({placeholders})
            """
            # print("Query SQL:", query) # Debug log

            cur.execute(query, list(dados.values()))
            
            conn.commit()
            # print("Dados inseridos com sucesso!") # Debug log
            
            # flash('Formulário enviado com sucesso! Aguarde a aprovação do administrador.', 'success')
            return redirect(url_for('dashboard_terapeuta'))
                
        except Exception as e:
            print(f"Erro detalhado: {str(e)}")  # Log mais detalhado do erro
            if conn:
                conn.rollback()
            flash('Erro ao enviar formulário. Por favor, tente novamente.', 'error')
            
        finally:
            if conn:
                cur.close()
                conn.close()
    
    return render_template('formulario_terapeuta.html', form=form)

def get_pg_placeholder(value):
    if isinstance(value, bool):
        return 'CAST(%s AS TINYINT)'
    elif isinstance(value, int):
        return 'CAST(%s AS INTEGER)'
        # return '%s::integer'
    elif isinstance(value, float):
        return 'CAST(%s AS FLOAT)'
        # return '%s::float'
    elif value is None:
        return 'CAST(%s AS TEXT)'
        # return '%s::text'
    else:
        return 'CAST(%s AS TEXT)'
        # return '%s::text'

@app.before_request
def log_request_info():
    print('Headers: %s', request.headers)
    print('Body: %s', request.get_data())

@app.route('/gerar_excel_reprovados_terapeutas')
@login_required
def gerar_excel_reprovados_terapeutas():
    if not current_user.is_authenticated or not current_user.is_admin():
        flash('Acesso não autorizado!', 'error')
        return redirect(url_for('admin_usuarios'))

    conn = conectar_bd()
    cur = conn.cursor()

    try:
        # Consulta os registros reprovados
        cur.execute("""
            SELECT 
                data_cadastro, nome_completo, cpf, telefone, 
                celular, email, cidade, estado, 
                endereco_consultorio, nivel_atual, ano_conclusao_avancado2, 
                ano_conclusao_sep, professores_formacao, formacao_academica, 
                participa_grupo_estudo, numero_supervisoes_ultimo_ano, 
                modalidade, faixa_valor_sessao, consultorio_acessivel, 
                observacao_acessibilidade, interesse_producao_cientifica, 
                associado_abt, sugestoes, concordou_termos, status
            FROM terapeuta_napese 
            WHERE status = 'reprovado'
            ORDER BY data_cadastro DESC
        """)
        
        registros = cur.fetchall()
        
        # if not registros:
            # flash('Não há registros reprovados para exportar!', 'info')
            # return redirect(url_for('admin_usuarios'))

        # Criar workbook e definir cabeçalhos
        wb = Workbook()
        ws = wb.active
        ws.title = "Terapeutas Reprovados"

        headers = [
            "Data Cadastro", "Nome Completo", "CPF", "Telefone", 
            "Celular", "Email", "Cidade", "Estado", 
            "Endereço Consultório", "Nível Atual", "Ano Conclusão Avançado 2",
            "Ano Conclusão SEP", "Professores Formação", "Formação Acadêmica",
            "Participa Grupo Estudo", "Número Supervisões", 
            "Modalidade", "Faixa Valor Sessão", "Consultório Acessível",
            "Observação Acessibilidade", "Interesse Produção Científica",
            "Associado ABT", "Sugestões", "Concordou Termos", "Status"
        ]

        # Estilizar cabeçalhos
        header_fill = PatternFill(start_color="96d232", end_color="96d232", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Adicionar dados
        for row, registro in enumerate(registros, 2):
            for col, valor in enumerate(registro, 1):
                if isinstance(valor, (datetime.date, datetime.datetime)):
                    valor = valor.strftime("%d/%m/%Y")
                elif isinstance(valor, bool):
                    valor = "Sim" if valor else "Não"
                ws.cell(row=row, column=col, value=valor)

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        # Salvar em memória e retornar
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return Response(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment;filename=terapeutas_reprovados_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )

    except Exception as e:
        print(f"Erro ao gerar Excel: {e}")
        flash('Erro ao gerar arquivo Excel!', 'error')
        return redirect(url_for('admin_usuarios'))

    finally:
        cur.close()
        conn.close()

@app.route('/gerar_excel_reprovados_pacientes')
@login_required
def gerar_excel_reprovados_pacientes():
    if not current_user.is_authenticated or not current_user.is_admin():
        flash('Acesso não autorizado!', 'error')
        return redirect(url_for('admin_usuarios'))

    conn = conectar_bd()
    cur = conn.cursor()

    try:
        # Consulta os registros reprovados
        cur.execute("""
            SELECT 
                data_cadastro, email, nome_completo, cpf, telefones,
                data_nascimento, cidade, estado, genero, profissao,
                preferencia_atendimento, renda_familiar, num_dependentes,
                sintomas_relevantes, medicacoes, substancias_psicoativas,
                historico_acidentes, historico_cirurgias, dores,
                acompanhamento_psiquiatrico, acompanhamento_psicologico,
                tecnicas_corporais, motivo_procura, vivenciou_trauma,
                descricao_evento, tempo_decorrido
            FROM formulario_napese 
            WHERE aprovado = 'reprovado'
            ORDER BY data_cadastro DESC
        """)
        
        registros = cur.fetchall()

        # if not registros:
            # flash('Não há registros reprovados para exportar!', 'info')
            # return redirect(url_for('admin_usuarios'))

        # Criar um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros Reprovados"

        # Definir cabeçalhos
        headers = [
            "Data Cadastro", "Email", "Nome Completo", "CPF", "Telefones",
            "Data Nascimento", "Cidade", "Estado", "Gênero", "Profissão",
            "Preferência Atendimento", "Renda Familiar", "Nº Dependentes",
            "Sintomas Relevantes", "Medicações", "Substâncias Psicoativas",
            "Histórico Acidentes", "Histórico Cirurgias", "Dores",
            "Acompanhamento Psiquiátrico", "Acompanhamento Psicológico",
            "Técnicas Corporais", "Motivo Procura", "Vivenciou Trauma",
            "Descrição Evento", "Tempo Decorrido"
        ]

        # Estilizar cabeçalhos
        header_fill = PatternFill(start_color="96d232", end_color="96d232", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Adicionar cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Adicionar dados
        for row, registro in enumerate(registros, 2):
            for col, valor in enumerate(registro, 1):
                # Formatar datas
                if isinstance(valor, (datetime.date, datetime.datetime)):
                    valor = valor.strftime("%d/%m/%Y")
                # Formatar booleanos
                elif isinstance(valor, bool):
                    valor = "Sim" if valor else "Não"
                
                ws.cell(row=row, column=col, value=valor)

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Criar o arquivo em memória
        from io import BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Nome do arquivo com data atual
        hoje = datetime.datetime.now().strftime("%Y%m%d")
        filename = f"pacientes_reprovados_{hoje}.xlsx"

        # Adicionar headers corretos na resposta
        return Response(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                "Content-Disposition": f"attachment;filename={filename}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )

    except Exception as e:
        print(f"Erro ao gerar Excel: {e}")
        flash('Erro ao gerar arquivo Excel!', 'error')
        return redirect(url_for('admin_usuarios'))

    finally:
        cur.close()
        conn.close()

@app.route('/termos-condicoes')
def termos_condicoes():
    return render_template('termos_condicoes.html')

@app.route('/uploads/cartas_recomendacao/<path:filename>')
@login_required  # Protege o acesso aos arquivos
def cartas_recomendacao(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/uploads/comprovantes_sessoes/<path:filename>')
@login_required  # Protege o acesso aos arquivos
def comprovantes_sessoes(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER_COMPROVANTE'], filename)

# Adicione no final do arquivo
if __name__ == '__main__':
    init_db()
    criar_usuarios_teste()
    app.run(debug=True)